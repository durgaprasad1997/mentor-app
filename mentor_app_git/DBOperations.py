
import os,sys,django,openpyxl
import click
import MySQLdb
from onlineproject import settings
from openpyxl import load_workbook




os.environ.setdefault("DJANGO_SETTINGS_MODULE", 'onlineproject.settings')
django.setup()



from onlineapp.models import Colleges
from onlineapp.models import Student
from onlineapp.models import Marks



@click.group()
@click.pass_context
def cli(ctx):


    schemaName = settings.DATABASES['default']['NAME']
    print(schemaName)
    print("program started")



@cli.command()
@click.pass_context
def cleardb(ctx):
    print("clearing daata")
    Student.objects.all().delete()
    Colleges.objects.all().delete()
    Marks.objects.all().delete()






@cli.command()
@click.pass_context
def createdb(ctx):
    print("clearing daata")


    schemaName = settings.DATABASES['default']['NAME']
    db = MySQLdb.connect("localhost", "root", "mysql")
    cursor = db.cursor()
    q = "CREATE SCHEMA IF NOT EXISTS "+schemaName
    cursor.execute(q)


    os.system("python manage.py makemigrations")
    os.system("python manage.py migrate")



@cli.command()
@click.pass_context
def dropdb(ctx):

    print("dropping db")
    schemaName = settings.DATABASES['default']['NAME']
    db = MySQLdb.connect("localhost", "root", "mysql",schemaName )
    cursor = db.cursor()
    q = """DROP DATABASE `onlinedb`;"""
    cursor.execute(q)

    db.commit()



@cli.command()
@click.pass_context
def populatedb(ctx):
    print("loading daata")
    workbook = load_workbook('C:/PythonCourse/students.xlsx')


    worksheet = workbook['Colleges']
    first = 0
    for row in worksheet.iter_rows():
        s1 = row[0].value
        s2 = row[1].value
        s3 = row[2].value
        s4 = row[3].value

        if first == 0:
            first = 1
        else:
            if (s1 != None):
                c = Colleges(Name=str(s1), Acronym=str(s2), Location=str(s3), Contact=str(s4))
                c.save()





    worksheet = workbook.get_sheet_by_name('Current')
    first=0
    for row in worksheet.iter_rows():
        s1 = row[0].value
        s2 = row[1].value
        s3 = row[2].value
        s4 = row[3].value
        s4 = s4.lower()
        #print(s1,s2,s3,s4)
        if(first==0):
            first=1
        else:
            if(s1 !=None ):
                temp=Colleges.objects.get(Acronym=str(s2))
                s=Student(Name=str(s1),Colleges=temp,Email_id=str(s3),Dbnames=str(s4),dropped_out=False)
                s.save()



    worksheet = workbook.get_sheet_by_name('Deletions')
    first = 0
    for row in worksheet.iter_rows():
        s1 = row[0].value
        s2 = row[1].value
        s3 = row[2].value
        s4 = row[3].value
        s4 = s4.lower()
        # print(s1,s2,s3,s4)
        if (first == 0):
            first = 1
        else:
            if (s1 != None):
                temp = Colleges.objects.get(Acronym=str(s2))
                s = Student(Name=str(s1), Colleges=temp, Email_id=str(s3), Dbnames=str(s4), dropped_out=True)
                s.save()







    workbook = load_workbook('C:/PythonCourse/summer_pre_assign3_output.xlsx')
    worksheet = workbook.get_sheet_by_name('Marks')


    i = 0
    for row in worksheet.iter_rows():

        if (i != 0):

            x = row[0].value
            t = x.split('_')

            s1 = t[2].lower()
            s2 = t[1]
            s3 = row[1].value
            s4 = row[2].value
            s5 = row[3].value
            s6 = row[4].value
            s7 = row[5].value
            try:
                col=Student.objects.get(Dbnames=s1)
                m=Marks( student=col,problem1=s3,problem2=s4,problem3=s5,problem4=s6,Total=s7)
                m.save()
            except Exception as e:
                print("some exception")

        else:
            i = 1













if __name__=='__main__':

    cli(obj={})


    # manager = Colleges.objects
    # querySets = Colleges.objects.all()
    # print(querySets)
    # for querySet in querySets:
    #     print(querySet)

